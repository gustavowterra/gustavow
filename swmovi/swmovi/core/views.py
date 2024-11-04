

from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import *
from django.conf import settings
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
import openpyxl


@csrf_exempt
def dologin(request):
    data = {}

    username = request.POST.get('user', '')  # Obtém o valor da chave 'username' ou retorna uma string vazia
    password = request.POST.get('pass', '')
    print(username, password)
    user = authenticate(username=username, password=password)
    print (user)
    if user is not None:
        login(request, user)
        return redirect('/painel/')

    else:
        data['msg'] = ' Usuario ou Senha Invalido(s)'
        data['class'] = 'alert-danger'

        return render(request, 'login/index.html', data)


@csrf_exempt
def painel(request):



    user = request.user.username

    if user:



        return render(request, 'modelo/index.html')
    else:
        return render(request, 'login/index.html')

def cadastro_estoque(request):

    return render (request, 'uploads/cadastro_em_lote.html')

def upload_estoque(request):

    if request.method == "POST":
        arquivo = request.FILES.get('arquivo')

        if not arquivo:
            return JsonResponse({"success": False, "message": "Nenhum arquivo foi enviado."}, status=400)

        elif arquivo:
            wb = openpyxl.load_workbook(arquivo)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):

                M_CO, M_OR, M_DE, M_NP, M_UM, M_UT, M_FA, M_MA, M_TI, M_GC, M_GE, M_GR, M_KMP, M_NAT= row
                if M_CO and M_DE:

                    Almoxarifado_Material.objects.create(
                        codigo=M_CO,
                        codigo_original=M_OR,
                        descricao=M_DE,
                        nome_praca=M_NP,
                        codigo_unidade_medida=M_UM,
                        utilizacao=M_UT,
                        codigo_fabricante=M_FA,
                        codigo_marca=M_MA,
                        codigo_tipo_condicao=M_TI,
                        codigo_grupo_contabil=M_GC,
                        codigo_grupo_estoque=M_GE,
                        codigo_grupo_compras=M_GR,
                        flag_km=M_KMP,
                        codigo_natureza=M_NAT,

                    )
                else:
                    print('Erro em algum dos campos necessarios')
                    return JsonResponse({"success": False, "message": "Faltam Informações"}, status=405)
                    pass

            return JsonResponse({"success": True, "message": "Dados importados com sucesso!"})

        else:
            return JsonResponse({"success": False, "message": f"Erro ao processar o arquivo: {str(e)}"}, status=500)

    return JsonResponse({"success": False, "message": "Método não permitido."}, status=405)

def cadastro_estoque_grupos(request):

    return render (request, 'uploads/cadastro_grupos_em_lote.html')
def upload_estoque_grupos(request):
    if request.method == "POST":
        arquivo = request.FILES.get('arquivo')

        if not arquivo:
            return JsonResponse({"success": False, "message": "Nenhum arquivo foi enviado."}, status=400)

        elif arquivo:
            wb = openpyxl.load_workbook(arquivo)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                MTM_CO, MTM_DE, MTM_TI = row
                if MTM_CO and MTM_DE:
                    tipo_id = MTM_TI
                    tipo_grupo = Grupos_Tipos.objects.filter(id=tipo_id).first()

                    Grupos_Manutencao.objects.create(
                        codigo=MTM_CO,
                        descricao=MTM_DE,
                        tipo=tipo_grupo
                    )
                else:
                    print('Erro em algum dos campos necessarios')
                    return JsonResponse({"success": False, "message": "Faltam Informações"}, status=405)
                    pass

            return JsonResponse({"success": True, "message": "Dados importados com sucesso!"})

        else:
            return JsonResponse({"success": False, "message": f"Erro ao processar o arquivo: {str(e)}"}, status=500)

    return JsonResponse({"success": False, "message": "Método não permitido."}, status=405)

def estoque_grupos_pecas (request):


    context = {

        'tipos': Grupos_Tipos.objects.all(),
        'grupos': Grupos_Manutencao.objects.all()


    }
    if request.method == 'POST':

        codigo = request.POST.get('codigo_grupo')
        descricao = request.POST.get('nome_grupo')
        tipo_grupo_id = request.POST.get('tipo_grupo')
        tipo_grupo = Grupos_Tipos.objects.filter(id=tipo_grupo_id).first()
        novo_grupo = Grupos_Manutencao()
        novo_grupo.tipo = tipo_grupo
        novo_grupo.descricao = descricao
        novo_grupo.codigo = codigo
        novo_grupo.save()

    return render (request, 'estoque/grupos.html', context )